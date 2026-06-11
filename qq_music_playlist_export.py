# --- QQ音乐歌单导出工具 ---
# -*- coding: utf-8 -*-

"""
依赖（自行安装）：
    pip install requests
    pip install openpyxl
"""

__author__ = "lengxiQwQ, Dragon水魅"

import os
import re
import sys
import json
import csv
import platform
import subprocess
import requests

# --- 辅助函数 ---

# 去掉可能的 JSONP 包装，返回 JSON 字符串
def strip_jsonp(text):
    m = re.search(r'^[^\(]*\((\{.*\})\)\s*;?\s*$', text, flags=re.S)
    if m:
        return m.group(1)
    return text

# 将文件名中 Windows 不允许的字符替换为空格：< > : " / \ | ? * 以及控制字符（0-31），去掉首尾空白
def sanitize_filename(name):
    if not name:
        return ""
    name = re.sub(r'[\x00-\x1f]', ' ', name)
    name = re.sub(r'[<>:"/\\|?*]', ' ', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name or "playlist"

# 从用户输入中提取歌单ID
def extract_playlist_id(text):
    if not text:
        return None
    text = text.strip()
    if re.fullmatch(r'\d+', text):
        return text
    m = re.search(r'(\d{5,})', text)
    if m:
        return m.group(1)
    return None

# --- 与 QQ音乐接口交互（返回 (playlist_title, [(name,singers,album), ...], author) 或 None） ---

# 获取指定 QQ号 的所有歌单列表，返回 (nickname, [{"id":..., "name":...}, ...]) 或 None
def get_user_playlists(uin):
    """
    调用 QQ音乐接口获取指定用户（QQ号）的所有歌单。
    返回 (nickname字符串, [{"id":歌单ID, "name":歌单名称}, ...]) 或 None（失败时）。
    """
    # 修复 QQ音乐 API 返回的双重编码字符串（UTF-8 → Latin-1 解码的乱码）
    def fix_encoding(s):
        if not isinstance(s, str):
            return str(s or "")
        try:
            # 尝试用 latin-1 编码回字节，再以 utf-8 解码修复乱码
            fixed = s.encode('latin-1').decode('utf-8')
            # 如果修复后有无效字符，回退原字符串
            return fixed
        except (UnicodeEncodeError, UnicodeDecodeError, LookupError):
            return s

    url = "https://c.y.qq.com/rsc/fcgi-bin/fcg_user_created_diss"
    params = {
        "hostUin": 0,
        "hostuin": str(uin),
        "sin": 0,
        "size": 200,
        "g_tk": 5381,
        "loginUin": 0,
        "format": "json",
        "inCharset": "utf8",
        "outCharset": "utf-8",
        "notice": 0,
        "platform": "yqq.json",
        "needNewCode": 0,
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://y.qq.com/portal/profile.html",
    }
    try:
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        j = resp.json()
        if isinstance(j, dict) and j.get("code") == 0:
            data = j.get("data") or {}
            # 提取昵称（hostname），修复编码
            nickname_raw = data.get("hostname") or ""
            nickname = fix_encoding(nickname_raw).strip() or str(uin)
            # 提取歌单列表（tid 是歌单ID，diss_name 是歌单名称）
            items = data.get("disslist") or []
            playlists = []
            for item in items:
                if not isinstance(item, dict):
                    continue
                pid = str(item.get("tid") or "")
                pname = fix_encoding(item.get("diss_name") or "")
                # 过滤掉 dir_show=0 的（如 QZone背景音乐）和 tid=0 的
                if pid and pid != "0" and item.get("dir_show", 1) != 0:
                    playlists.append({"id": pid, "name": pname})
            if playlists:
                return (nickname, playlists)
    except Exception:
        pass
    
    return None

# 老接口: c.y.qq.com 获取歌单信息 返回 (title, song_list, author) 或 None
def try_c_y_qq(disstid):
    url = "https://c.y.qq.com/qzone/fcg-bin/fcg_ucc_getcdinfo_byids_cp.fcg"
    params = {
        "disstid": str(disstid),
        "type": "1",
        "json": "1",
        "utf8": "1",
        "onlysong": "0",
        "format": "json"
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"https://y.qq.com/n/ryqq/playlist/{disstid}"
    }

    try:
        resp = requests.get(url, params=params, headers=headers, timeout=10)
    except Exception:
        return None

    text = strip_jsonp(resp.text.strip())
    try:
        j = json.loads(text)
    except Exception:
        return None

    title = None
    author = ""
    if isinstance(j, dict):
        if "cdlist" in j and isinstance(j["cdlist"], list) and j["cdlist"]:
            cd = j["cdlist"][0]
            title = cd.get("dissname") or cd.get("diss_name") or cd.get("title") or cd.get("cdtitle") or cd.get("name")
            author = cd.get("data_signer") or cd.get("nickname") or cd.get("nick") or cd.get("username") or cd.get("uname") or ""
            songlist = cd.get("songlist") or cd.get("list") or []
        else:
            title = j.get("dissname") or j.get("title") or j.get("name")
            author = j.get("data_signer") or j.get("nickname") or j.get("nick") or ""
            songlist = j.get("songlist") or []
    else:
        return None

    results = []
    for s in songlist:
        name = s.get("songname") or s.get("name") or s.get("title") or ""
        if "singer" in s and isinstance(s["singer"], list):
            singers = ", ".join([a.get("name") or a.get("singer_name") or a.get("nickname","") for a in s["singer"]])
        else:
            singers = s.get("singername") or s.get("singer_name") or s.get("lan") or s.get("singer") or ""
        album = s.get("albumname") or (s.get("album") or {}).get("name") or s.get("albumname_utf8") or ""
        results.append((name, singers, album))

    return (title or "", results, author)

# 新接口: u.y.qq.com 的 GetPlaylistDetail 返回 (title, song_list, author) 或 None
def try_u_y_qq_playlist_detail(playlist_id):
    url = "https://u.y.qq.com/cgi-bin/musicu.fcg"
    payload = {
        "comm": {
            "cv": 4747474,
            "ct": 24,
            "format": "json",
            "inCharset": "utf-8",
            "outCharset": "utf-8",
            "notice": 0,
            "platform": "yqq.json",
            "needNewCode": 1,
            "uin": "0"
        },
        "playlist": {
            "method": "GetPlaylistDetail",
            "module": "music.playlist.PlaylistDetailServer",
            "param": {
                "id": int(playlist_id),
                "n": 1000,
                "order": 5
            }
        }
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"https://y.qq.com/n/ryqq/playlist/{playlist_id}",
        "Content-Type": "application/json"
    }
    try:
        resp = requests.post(url, data=json.dumps(payload), headers=headers, timeout=10)
    except Exception:
        return None

    try:
        j = resp.json()
    except Exception:
        return None

    pl = j.get("playlist")
    if not pl:
        return None

    title = ""
    author = ""
    if isinstance(pl, dict):
        title = pl.get("title") or pl.get("dissname") or (pl.get("data") or {}).get("title") or (pl.get("data") or {}).get("name") or pl.get("name") or ""
        creator = pl.get("creator") or (pl.get("data") or {}).get("creator") or {}
        if isinstance(creator, dict):

            # 兼容各种字段名
            author = creator.get("nickname") or creator.get("nickName") or creator.get("nick") or creator.get("name") or creator.get("data_signer") or ""
        if not author:
            author = pl.get("data_signer") or pl.get("nickname") or pl.get("nick") or ""

    if pl.get("code") is not None and pl.get("code") != 0:
        return None

    songlist = (pl.get("data") or {}).get("songlist") or (pl.get("data") or {}).get("songs") or pl.get("songlist") or []
    results = []
    for s in songlist:
        name = s.get("name") or s.get("songname") or ""
        if isinstance(s.get("singer"), list):
            singers = ", ".join([a.get("name","") for a in s["singer"]])
        else:
            singers = s.get("singername") or s.get("singer") or ""
        album = (s.get("album") or {}).get("name") or s.get("albumname") or ""
        results.append((name, singers, album))
    return (title or "", results, author)

def get_playlist_songs(playlist_id):

    # 依次尝试不同接口，返回 (title, songs, author) 或 None
    res = try_c_y_qq(playlist_id)
    if res:
        return res

    res = try_u_y_qq_playlist_detail(playlist_id)
    if res:
        return res

    # 备用 payload（playlist_songlist），尝试从 data 中提取 author
    url = "https://u.y.qq.com/cgi-bin/musicu.fcg"
    payload_alt = {
        "comm": {"cv": 4747474, "ct": 24, "format": "json", "uin": "0", "platform": "yqq.json"},
        "playlist_songlist": {
            "method": "GetPlaylistSongs",
            "module": "playlist.PlaylistSongListSrv",
            "param": {"id": int(playlist_id), "start": 0, "count": 1000}
        }
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"https://y.qq.com/n/ryqq/playlist/{playlist_id}",
        "Content-Type": "application/json"
    }
    try:
        resp = requests.post(url, data=json.dumps(payload_alt), headers=headers, timeout=10)
        j = resp.json()
        p = j.get("playlist_songlist", {})
        if p.get("code") == 0:
            data = p.get("data", {})
            songlist = data.get("songlist", []) or []
            title = data.get("title") or data.get("dissname") or data.get("name") or ""
            results = []
            for s in songlist:
                name = s.get("name","")
                singers = ", ".join([a.get("name","") for a in s.get("singer", [])]) if isinstance(s.get("singer"), list) else s.get("singer","")
                album = (s.get("album") or {}).get("name","")
                results.append((name, singers, album))
            author = data.get("data_signer") or data.get("nickname") or data.get("nick") or data.get("username") or ""
            return (title or "", results, author)
    except Exception:
        pass

    return None

# --- 导出函数（txt/csv/xlsx/json） ---

# 保存为 txt（每行 "Title - Artist - Album"）
def export_to_txt(rows, output_path):
    header = ""
    try:
        with open(output_path, "w", encoding="utf-8-sig", newline="\n") as f:
            f.write(header)
            for name, singers, album in rows:
                safe_name = (name or "").replace("\r"," ").replace("\n"," ")
                safe_singers = (singers or "").replace("\r"," ").replace("\n"," ")
                safe_album = (album or "").replace("\r"," ").replace("\n"," ")
                f.write(f"{safe_name} - {safe_singers} - {safe_album}\n")
    except Exception:
        raise

# 保存为 csv（utf-8-sig，第一行为表头）
def export_to_csv(rows, output_path):
    try:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Title", "Artist", "Album"])
            for name, singers, album in rows:
                safe_name = (name or "").replace("\r"," ").replace("\n"," ")
                safe_singers = (singers or "").replace("\r"," ").replace("\n"," ")
                safe_album = (album or "").replace("\r"," ").replace("\n"," ")
                writer.writerow([safe_name, safe_singers, safe_album])
    except Exception:
        raise

# 保存为 xlsx（openpyxl），第一行为表头
def export_to_xlsx(rows, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise

    wb = Workbook()
    ws = wb.active
    ws.title = "QQ音乐歌单"
    ws.append(["Title", "Artist", "Album"])
    for name, singers, album in rows:
        ws.append([name or "", singers or "", album or ""])
    for col_idx in range(1, 4):
        col = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col]:
            try:
                l = len(str(cell.value))
            except:
                l = 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[col].width = min(max(10, int(max_len * 1.1) + 2), 60)
    wb.save(output_path)

# 保存为 json（数组对象），使用 utf-8 编码
def export_to_json(rows, output_path):
    data = []
    for name, singers, album in rows:
        data.append({
            "Title": name or "",
            "Artist": singers or "",
            "Album": album or ""
        })
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        raise

# 打开并选中文件（Windows），或在其他平台打开文件所在目录
def open_file_location(path):
    try:
        abs_path = os.path.abspath(path)
        dirpath = os.path.dirname(abs_path)
        system = platform.system()
        if system == "Windows":

            # explorer /select, 可以高亮选中文件
            try:
                subprocess.run(["explorer", "/select,", abs_path])
                return
            except Exception:

                # 退回到打开目录
                try:
                    os.startfile(dirpath)
                    return
                except Exception:
                    pass
        elif system == "Darwin":

            # macOS
            try:
                subprocess.run(["open", dirpath])
                return
            except Exception:
                pass
        else:

            # Linux / 其他，尝试 xdg-open
            try:
                subprocess.run(["xdg-open", dirpath])
                return
            except Exception:
                pass
    except Exception:
        pass

# --- 批量导出共享函数 ---
def batch_export_songs(playlists, folder_name, nickname=""):
    """批量导出歌单列表"""
    print("\n请选择导出格式（将应用于所有歌单）：")
    print("  1) .xlsx  - Excel 文件")
    print("  2) .csv   - 标准 CSV utf-8-sig")
    print("  3) .json  - JSON 文件")
    print("  4) .txt   - 纯文本格式")
    fmt_choice = input("选择 (1-4)：").strip() or "1"
    if fmt_choice not in ("1","2","3","4"):
        print("选择无效，默认使用 xlsx")
        fmt_choice = "1"
    # 创建文件夹
    os.makedirs(folder_name, exist_ok=True)
    print(f"\n开始批量导出，保存到文件夹：{folder_name}")
    exported = 0
    failed = 0
    for idx, pl in enumerate(playlists):
        pid = pl["id"]
        pname = pl["name"]
        print(f"  [{idx+1}/{len(playlists)}] 正在获取歌单 [{pid}] {pname}...")
        res = get_playlist_songs(pid)
        if not res or not res[1]:
            print(f"    ✗ 跳过（无法读取或歌单为空）")
            failed += 1
            continue
        title, songs, author = res
        safe_title = sanitize_filename(title or pname)
        safe_author = sanitize_filename(author or nickname or "未知")
        out_path = None
        try:
            if fmt_choice == "1":
                out_path = os.path.join(folder_name, f"{safe_title} - {safe_author}.xlsx")
                export_to_xlsx(songs, out_path)
            elif fmt_choice == "2":
                out_path = os.path.join(folder_name, f"{safe_title} - {safe_author}.csv")
                export_to_csv(songs, out_path)
            elif fmt_choice == "3":
                out_path = os.path.join(folder_name, f"{safe_title} - {safe_author}.json")
                export_to_json(songs, out_path)
            elif fmt_choice == "4":
                out_path = os.path.join(folder_name, f"{safe_title} - {safe_author}.txt")
                export_to_txt(songs, out_path)
            print(f"    ✓ 已导出（{len(songs)} 首）：{os.path.basename(out_path)}")
            exported += 1
        except ImportError:
            print(f"    ✗ 导出失败：缺少 openpyxl 库，请运行：pip install openpyxl")
            failed += 1
        except Exception as e:
            print(f"    ✗ 导出失败：{e}")
            failed += 1
    print(f"\n{'='*40}")
    print(f"批量导出完成！成功：{exported}，失败：{failed}")
    abs_folder = os.path.abspath(folder_name)
    print(f"文件保存在：{abs_folder}")
    try:
        if platform.system() == "Windows":
            os.startfile(abs_folder)
        elif platform.system() == "Darwin":
            subprocess.run(["open", abs_folder])
        else:
            subprocess.run(["xdg-open", abs_folder])
    except Exception:
        pass
    return (exported, failed)

# --- 主程序入口 ---
def main():
    print("============ QQ音乐歌单导出工具 ============")
    try:
        while True:
            user_input = input("\n请输入歌单链接 或 歌单ID 或 QQ号（输入 0 退出）：").strip()
            if user_input in ("0", "q", "quit", "exit"):
                print("\n========== 程序已退出，感谢使用！===========")
                break

            is_just_digits = user_input.isdigit() and len(user_input) >= 4
            pid = extract_playlist_id(user_input)

            # 同时尝试歌单ID和QQ号
            playlist_info = None
            qq_info = None

            if pid:
                playlist_info = get_playlist_songs(pid)
            
            if is_just_digits:
                qq_info = get_user_playlists(user_input)

            # ── 既是歌单又是QQ号，让用户选择 ──
            if playlist_info and playlist_info[1] and qq_info:
                title, songs, author = playlist_info
                qq_nickname, qq_playlists = qq_info
                print(f"\n'{user_input}' 既是歌单ID也是QQ号，请选择：")
                print(f"  1) 导出该歌单 — [{pid}] {title}（{len(songs)} 首）")
                print(f"  2) 导出该QQ号所有歌单 — {qq_nickname} 共 {len(qq_playlists)} 个歌单")
                choice = input("请选择 (1/2，默认1)：").strip()
                if choice == "2":
                    print(f"\n用户：{qq_nickname}，共 {len(qq_playlists)} 个歌单：")
                    for i, pl in enumerate(qq_playlists):
                        print(f"  {i+1:3d}. [{pl['id']}] {pl['name']}")
                    batch_export_songs(qq_playlists, user_input, qq_nickname)
                    continue
                # 否则走单歌单导出

            # ── 只是QQ号 ──
            elif qq_info:
                qq_nickname, qq_playlists = qq_info
                print(f"\n检测到QQ号，用户：{qq_nickname}，共 {len(qq_playlists)} 个歌单：")
                for i, pl in enumerate(qq_playlists):
                    print(f"  {i+1:3d}. [{pl['id']}] {pl['name']}")
                batch_export_songs(qq_playlists, user_input, qq_nickname)
                continue

            # ── 都不匹配 ──
            elif not playlist_info and not qq_info:
                print("\n无法识别输入，请确认歌单链接/ID 或 QQ号是否正确")
                continue

            # ── 单歌单导出流程 ──
            title, songs, author = playlist_info
            pid = pid or extract_playlist_id(user_input)

            playlist_title = title.strip() if title else f"playlist_{pid}"
            safe_title = sanitize_filename(playlist_title)
            display_author = author or "未知"
            safe_author = sanitize_filename(display_author) or "未知作者"

            print(f"\n已获取到 {display_author} 的歌单：\n名称：{playlist_title}，共 {len(songs)} 首歌曲")
            print("=" * 42)
            print("请选择导出格式：")
            print(" 1) .xlsx  - (默认) Excel 文件")
            print(" 2) .csv   - 标准 CSV utf-8-sig")
            print(" 3) .json  - JSON 文件，数组")
            print(" 4) .txt   - 纯文本格式")
            print("=" * 42)
            choice = input("选择 (1-4，输入 0 退出程序)：").strip() or "1"
            if choice == "0":
                print("\n========== 程序已退出，感谢使用！===========")
                break
            if choice not in ("1","2","3","4"):
                print("选择无效，默认使用 xlsx")
                choice = "1"

            try:
                out_name = None
                if choice == "1":
                    out_name = f"{safe_title} - {safe_author}.xlsx"
                    try:
                        export_to_xlsx(songs, out_name)
                    except ImportError:
                        print("导出 xlsx 失败：缺少 openpyxl 库，请运行：pip install openpyxl")
                        out_name = None
                elif choice == "2":
                    out_name = f"{safe_title} - {safe_author}.csv"
                    export_to_csv(songs, out_name)
                elif choice == "3":
                    out_name = f"{safe_title} - {safe_author}.json"
                    export_to_json(songs, out_name)
                elif choice == "4":
                    out_name = f"{safe_title} - {safe_author}.txt"
                    export_to_txt(songs, out_name)
                
                if out_name:
                    print(f"\n已保存为: {out_name}")
                    print("=" * 42)
                    open_file_location(out_name)
            except Exception as e:
                print("保存文件时出错：", e)
    except KeyboardInterrupt:
        print("\n\n========= 程序被用户终止，感谢使用！=========")

if __name__ == "__main__":
    main()
